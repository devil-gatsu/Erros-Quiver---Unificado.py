import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import datetime
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# -------------------------------------------------------------
# Configuração Global do Tema (NextGen)
# -------------------------------------------------------------
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ------------------------------
# Utilidades de UI (thread-safe)
# ------------------------------
def set_progress_safe(root, canvas, value):
    try:
        root.after(0, lambda: desenhar_foguete(canvas, value))
    except Exception:
        pass

def show_info_safe(root, title, msg):
    root.after(0, lambda: messagebox.showinfo(title, msg))

def show_error_safe(root, title, msg):
    root.after(0, lambda: messagebox.showerror(title, msg))

def update_status_safe(root, status_label, msg, text_color="#FFFFFF"):
    try:
        root.after(0, lambda: status_label.configure(text=msg, text_color=text_color))
    except Exception:
        pass

# ------------------------------
# Animação do Foguete (Loading)
# ------------------------------
def desenhar_foguete(canvas, percent):
    canvas.delete("all")
    w = canvas.winfo_width()
    h = canvas.winfo_height()
    if w <= 1: w = 500
    if h <= 1: h = 50
    
    # Pista Neon
    canvas.create_line(20, h/2, w - 20, h/2, fill="#2b2b2b", width=8, capstyle=tk.ROUND)
    
    # Progresso
    x_pos = 20 + (percent / 100.0) * (w - 60)
    if percent > 0:
        canvas.create_line(20, h/2, x_pos, h/2, fill="#00e5ff", width=8, capstyle=tk.ROUND)
        
    # Foguete e Gás
    if percent > 0 and percent < 100:
        canvas.create_text(x_pos - 15, h/2, text="💨", font=("Segoe UI", 12))
    canvas.create_text(x_pos + 10, h/2, text="🚀", font=("Segoe UI", 18))

# -------------------------------------------------------------
# Formatação Premium do Excel
# -------------------------------------------------------------
def formatar_planilha_saida(caminho):
    try:
        wb = load_workbook(caminho)
        ws = wb.active
        
        fill_header = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        font_header = Font(color="00E5FF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = min(max(max_length, len(str(cell.value))), 60)
                except: pass
            ws.column_dimensions[column_letter].width = (max_length + 2)
            
        wb.save(caminho)
    except Exception as e:
        print(f"Erro ao formatar Excel: {e}")

# -------------------------------------------------------------
# 1) Garantias para Locação (LÓGICA RESTAURADA)
# -------------------------------------------------------------
def gpl_comparar_ponto_venda_vendedor(row):
    ponto_venda = str(row.get('PTO_NOME', '')).strip()
    vendedor = str(row.get('VENDEDOR', '')).strip()
    ignore_list = ['Jetimob', 'Renata guedes', 'Comercial totumseg', 'matriz', 'marco zero empreendimentos']
    if ponto_venda in ignore_list or vendedor in ignore_list: return ''
    if ponto_venda.lower().startswith(vendedor.lower()) or vendedor.lower().startswith(ponto_venda.lower()): return ''
    return f"Erro: Ponto de venda {ponto_venda} não coincide com vendedor {vendedor}"

def gpl_verificar_ausencia_vendedor(row):
    if pd.isna(row.get('VENDEDOR')) or str(row.get('VENDEDOR', '')).strip() == '': return "Erro: Vendedor ausente"
    return ''

def gpl_verificar_vigencia_zerada(row):
    try:
        if pd.to_numeric(row.get('FINAL_VIGENCIA'), errors='coerce') - pd.to_numeric(row.get('INICIO_VIGENCIA'), errors='coerce') == 0:
            return "Erro: Vigência zerada"
    except Exception as e: return f"Erro na vigência: {str(e)}"
    return ''

def gpl_verificar_produto_divergente(row):
    seguradora = str(row.get('SEGURADORA', '')).strip()
    produto = str(row.get('PRODUTO', '')).strip()
    produtos_aceitos = {
        'BrasilCap': ['Título de capitalização'],
        'MAPFRECAP': ['FIANCA LOCATICIA TAXA FIXA MAPFRE', 'FIANCA LOCATICIA MAPFRE CONNECT'],
        'Porto Capitalização': ['FIANCA LOCATICIA', 'FIANCA PORTO ESSENCIAL'],
        'ICATU': ['Título de capitalização'],
        'SUL AMERICA': ['Título de capitalização'],
        'Liberty': ['Fiança Locatícia', 'Fiança Liberty Online'],
        'Yelum': ['Fiança Locatícia', 'Fiança Liberty Online'],
    }
    if seguradora in produtos_aceitos and produto not in produtos_aceitos[seguradora]:
        return f"Erro: Produto {produto} divergente para a seguradora {seguradora}"
    return ''

def gpl_verificar_proposta_cia_vazia(row):
    if pd.isna(row.get('PROPOSTA_CIA')) and row.get('TIPO_DOCUMENTO') != 'cancelamento' and row.get('QTDE_SINISTROS_DOCUMENTO', 0) != 0:
        return "Erro: Proposta_CIA vazia"
    return ''

def gpl_verificar_numero_arquivo_vazio(row):
    if pd.isna(row.get('NUMERO_ARQUIVO')) or str(row.get('NUMERO_ARQUIVO', '')).strip() == '':
        return "Erro: Número de arquivo vazio"
    return ''

def gpl_verificar_data_envio_post_emissao(row):
    try:
        formato = "%d/%m/%Y"
        data_envio = pd.to_datetime(row.get('DATA_ENVIO_PROPOSTA'), format=formato, errors='coerce')
        data_emissao = pd.to_datetime(row.get('DATA_EMISSAO'), format=formato, errors='coerce')
        if pd.notna(data_envio) and pd.notna(data_emissao) and data_envio.year >= 2023 and data_envio > data_emissao:
            return "Erro: Data de envio posterior à data de emissão"
    except Exception as e: return f"Erro na data: {str(e)}"
    return ''

def gpl_verificar_placa_matricula(row):
    placa_valida = ['PF RESIDENCIAL', 'PF NÃO RESIDENCIAL', 'PJ RESIDENCIAL', 'PJ NÃO RESIDENCIAL']
    val = str(row.get('PLACA_MATRICULA', '')).strip()
    if val == '': return "Erro: Placa da matrícula vazia"
    if val not in placa_valida: return f"Erro: Placa da matrícula {val} diferente do padrão"
    return ''

def gpl_verificar_ocupacao_vazia(row):
    if pd.isna(row.get('OCUPACAO')) or str(row.get('OCUPACAO', '')).strip() == '': return "Erro: Ocupação vazia"
    return ''

def gpl_verificar_descricao_bem_vazia(row):
    if pd.isna(row.get('DESCRICAO_BEM')) or str(row.get('DESCRICAO_BEM', '')).strip() == '': return "Erro: Descrição do bem vazia"
    return ''

def gpl_verificar_apolice_asterisco(row):
    return ''

def pipeline_garantias_locacao(df, root, canvas):
    total = len(df)
    for i, row in df.iterrows():
        erro = ' | '.join(filter(None, [
            gpl_comparar_ponto_venda_vendedor(row),
            gpl_verificar_ausencia_vendedor(row),
            gpl_verificar_vigencia_zerada(row),
            gpl_verificar_produto_divergente(row),
            gpl_verificar_proposta_cia_vazia(row),
            gpl_verificar_numero_arquivo_vazio(row),
            gpl_verificar_data_envio_post_emissao(row),
            gpl_verificar_placa_matricula(row),
            gpl_verificar_ocupacao_vazia(row),
            gpl_verificar_descricao_bem_vazia(row),
            gpl_verificar_apolice_asterisco(row),
        ]))
        df.at[i, 'Erro'] = erro if erro else "Erros não localizados"
        set_progress_safe(root, canvas, ((i + 1) / max(total, 1)) * 100)

    colunas_ordem = [
        'PTO_NOME', 'VENDEDOR', 'APOLICE', 'CLIENTE', 'INICIO_VIGENCIA', 'FINAL_VIGENCIA', 'SEGURADORA',
        'PRODUTO', 'PROPOSTA_CIA', 'NUMERO_ARQUIVO', 'ENDOSSO', 'TIPO_MOVIMENTO', 'TIPO_DOCUMENTO',
        'PREMIO_PARCELAS', 'PREMIO_TOTAL', 'QTDE_SINISTROS_DOCUMENTO', 'SITUACAO_APOLICE', 'DATA_EMISSAO',
        'QTD_PARCELAS', 'DATA_ENVIO_PROPOSTA', 'DATA_ENTRADA', 'DATA_PROPOSTA', 'VALOR_REPASSES',
        'PERCENTUAL_REPASSE', 'PLACA_MATRICULA', 'OCUPACAO', 'DESCRICAO_BEM', 'PERC_COMISSAO',
        'AUD_INCLUSAO_USR', 'AUD_ALTERACAO_USR', 'AUD_INCLUSAO_DATA', 'Erro'
    ]
    return df[[c for c in colunas_ordem if c in df.columns]]

# -------------------------------------------------------------
# 2) Seguros Imobiliários (LÓGICA RESTAURADA)
# -------------------------------------------------------------
def sim_comparar_ponto_venda_vendedor(row):
    ponto_venda = str(row.get('PTO_NOME', '')).strip()
    vendedor = str(row.get('VENDEDOR', '')).strip()
    ignore_list = ['Jetimob', 'Renata guedes', 'Comercial totumseg', 'matriz', 'marco zero empreendimentos', 'JETIMOB - ORUS TECNOLOGIA LTDA - ME']
    if ponto_venda in ignore_list or vendedor in ignore_list: return ''
    if vendedor == "Crédito Real" and (ponto_venda in ["Pontal", "SP ADM E LOCADORA"]): return ''
    if vendedor == "DEMBOSKI" and ponto_venda == "ANAGE": return ''
    if vendedor == "Jetimob" and ponto_venda: return ''
    if ponto_venda.lower().startswith(vendedor.lower()) or vendedor.lower().startswith(ponto_venda.lower()): return ''
    return f"Erro: Ponto de venda {ponto_venda} não coincide com vendedor {vendedor}"

def sim_verificar_proposta_cia_vazia(row):
    produtos_permitidos = ['VIDA', 'PRÓ-TRABALHO', 'VIDA INDIVIDUAL (PREMIO TOTAL MENSAL)', 'VIDA EM GRUPO', 'VIDA INDIVIDUAL (PREMIO INTEGRAL)']
    if pd.isna(row.get('PROPOSTA_CIA')) and row.get('PRODUTO') not in produtos_permitidos:
        return "Erro: Proposta_CIA vazia"
    return ''

def pipeline_seguros_imobiliarios(df, root, canvas):
    total = len(df)
    for i, row in df.iterrows():
        erro = ' | '.join(filter(None, [
            gpl_verificar_apolice_asterisco(row),
            sim_comparar_ponto_venda_vendedor(row),
            gpl_verificar_ausencia_vendedor(row),
            gpl_verificar_vigencia_zerada(row),
            sim_verificar_proposta_cia_vazia(row),
            gpl_verificar_numero_arquivo_vazio(row),
            gpl_verificar_data_envio_post_emissao(row),
        ]))
        df.at[i, 'Erro'] = erro if erro else "Erros não localizados"
        set_progress_safe(root, canvas, ((i + 1) / max(total, 1)) * 100)

    colunas_ordem = [
        'PTO_NOME', 'VENDEDOR', 'APOLICE', 'CLIENTE', 'INICIO_VIGENCIA', 'FINAL_VIGENCIA', 'SEGURADORA',
        'PRODUTO', 'PROPOSTA_CIA', 'NUMERO_ARQUIVO', 'ENDOSSO', 'TIPO_MOVIMENTO', 'TIPO_DOCUMENTO',
        'PREMIO_PARCELAS', 'PREMIO_TOTAL', 'QTDE_SINISTROS_DOCUMENTO', 'SITUACAO_APOLICE', 'DATA_EMISSAO',
        'QTD_PARCELAS', 'DATA_ENVIO_PROPOSTA', 'DATA_ENTRADA', 'DATA_PROPOSTA', 'VALOR_REPASSES',
        'PERCENTUAL_REPASSE', 'PERC_COMISSAO', 'AUD_INCLUSAO_USR', 'AUD_ALTERACAO_USR', 'AUD_INCLUSAO_DATA', 'Erro'
    ]
    return df[[c for c in colunas_ordem if c in df.columns]]

# -------------------------------------------------------------
# 3) Faturamento Educacional (LÓGICA RESTAURADA)
# -------------------------------------------------------------
def fe_verificar_vigencia_zerada(row):
    try:
        if pd.to_numeric(row.get('FINAL_VIGENCIA'), errors='coerce') - pd.to_numeric(row.get('INICIO_VIGENCIA'), errors='coerce') == 0:
            if row.get('TIPO_MOVIMENTO') in ['Cancelamento', 'Fatura Complementar', 'Não Emitida', 'Informativo']: return ''
            return "Erro: Vigência zerada"
    except Exception as e: return f"Erro na vigência: {str(e)}"
    return ''

def fe_verificar_seguradora(row):
    if str(row.get('SEGURADORA', '')).strip() not in ["MAPFRE SEG GERAIS", "MAPFRE VIDA S.A."]: return "Erro: Seguradora inválida"
    return ''

def fe_verificar_produto(row):
    if str(row.get('PRODUTO', '')).strip() not in ['2172 MAPFRE PROTEÇÃO ESCOLAR MULTIFLEX', '2173 MAPFRE PROTEÇÃO ESCOLAR MULTIFLEX', '2172 MAPFRE PROTEÇÃO EDUC MULTIFLEX']: return "Erro: Produto inválido"
    return ''

def fe_verificar_numero_arquivo(row):
    if pd.isna(row.get('NUMERO_ARQUIVO')) or str(row.get('NUMERO_ARQUIVO', '')).strip() == '':
        if pd.isna(row.get('PROPOSTA_CIA')) or str(row.get('PROPOSTA_CIA', '')).strip() == '':
            return "Erro: Número de arquivo vazio sem justificativa"
    return ''

def fe_verificar_endosso(row):
    if pd.isna(row.get('ENDOSSO')) or str(row.get('ENDOSSO', '')).strip() == '':
        if row.get('TIPO_MOVIMENTO') in ['Fatura', 'END. SEM. MOVIMENTO']:
            try:
                data_emissao = pd.to_datetime(row.get('DATA_EMISSAO'), format="%d/%m/%Y", errors='coerce')
                if pd.notna(data_emissao) and data_emissao > datetime.datetime.now() - datetime.timedelta(days=30): return ''
            except Exception: pass
            if pd.isna(row.get('PROPOSTA_CIA')) or str(row.get('PROPOSTA_CIA', '')).strip() == '':
                return "Erro: ENDOSSO vazio sem justificativa"
    endosso = str(row.get('ENDOSSO', '')).strip().lower()
    if 'cancelamento' in endosso and 'cancela' not in endosso: return "Erro: Problema de padronização no ENDOSSO"
    return ''

def fe_verificar_premio(row):
    if row.get('TIPO_MOVIMENTO') in ['Apólice', 'Cancelamento']:
        if pd.notna(row.get('PREMIO_PARCELAS')) or pd.notna(row.get('PREMIO_TOTAL')):
            return "Erro: PRÊMIO não deve ser preenchido para apólice ou cancelamento"
    return ''

def pipeline_faturamento_educacional(df, root, canvas):
    total = len(df)
    for i, row in df.iterrows():
        erro = ' | '.join(filter(None, [
            fe_verificar_vigencia_zerada(row),
            fe_verificar_seguradora(row),
            fe_verificar_produto(row),
            fe_verificar_numero_arquivo(row),
            fe_verificar_endosso(row),
            fe_verificar_premio(row),
            gpl_verificar_data_envio_post_emissao(row),
        ]))
        df.at[i, 'Error'] = erro if erro else 'Erros não localizados'
        set_progress_safe(root, canvas, ((i + 1) / max(total, 1)) * 100)

    colunas_ordem = [
        'PTO_NOME', 'APOLICE', 'CLIENTE', 'INICIO_VIGENCIA', 'FINAL_VIGENCIA', 'SEGURADORA', 'PRODUTO',
        'PROPOSTA_CIA', 'NUMERO_ARQUIVO', 'ENDOSSO', 'TIPO_MOVIMENTO', 'TIPO_DOCUMENTO', 'PREMIO_PARCELAS',
        'PREMIO_TOTAL', 'QTDE_SINISTROS_DOCUMENTO', 'AUD_INCLUSAO_DATA', 'SITUACAO_APOLICE', 'DATA_EMISSAO', 'QTD_PARCELAS',
        'DATA_ENVIO_PROPOSTA', 'DATA_ENTRADA', 'DATA_PROPOSTA', 'AUD_INCLUSAO_USR','AUD_ALTERACAO_USR',
        'VALOR_REPASSES', 'PERC_COMISSAO', 'Error'
    ]
    return df[[c for c in colunas_ordem if c in df.columns]]

# -------------------------------------------------------------
# 4) Sinistro Educacional (LÓGICA RESTAURADA)
# -------------------------------------------------------------
def sin_converte_data(data):
    try: return pd.to_datetime(data, format='%d/%m/%Y', errors='coerce')
    except Exception: return pd.NaT

def pipeline_sinistro_educacional(df, root, canvas):
    valid_columns = ['CLIENTE', 'CPF_CNPJ', 'SIN_NUMERO', 'SINSIT_DESCRICAO', 'BEM_SINISTRADO', 'CODIGO_CONTROLE',
                     'TIPO_SINISTRO', 'APOLICE', 'VALOR_RECLAMADO', 'VALOR_INDENIZADO', 'VALOR_LIBERADO',
                     'DATA_SINISTRO', 'DATA_AVISO', 'DATA_VISTORIA', 'DATA_LIBERACAO', 'DATA_LIQUIDACAO', 'LOCAL_SINISTRO']
    df = df[[c for c in valid_columns if c in df.columns]].copy()
    total = len(df)

    for i, row in df.iterrows():
        error_message = []
        codigo = row.get('CODIGO_CONTROLE')
        if pd.isna(codigo) or codigo == '': error_message.append("Código de controle vazio")
        else:
            if row.get('SINSIT_DESCRICAO') in ['LIQUIDADO', 'ANUIDADE EM LIQUIDAÇÃO', 'PERDA DE RENDA EM LIQUIDACAO']:
                if 'P' not in str(codigo) and 'T' not in str(codigo) and 'TRONADOR' not in str(row.get('LOCAL_SINISTRO', '')).upper():
                    error_message.append("Código de controle sem P ou T e não é TRONADOR")

        if (pd.isna(row.get('VALOR_RECLAMADO')) or row.get('VALOR_RECLAMADO') == 0) and (pd.isna(row.get('LOCAL_SINISTRO')) or row.get('LOCAL_SINISTRO') == ''):
            error_message.append("Sinistro sem valor reclamado e sem LOCAL_SINISTRO")

        try:
            d_sin = sin_converte_data(row.get('DATA_SINISTRO'))
            d_avi = sin_converte_data(row.get('DATA_AVISO'))
            d_vis = sin_converte_data(row.get('DATA_VISTORIA'))
            d_lib = sin_converte_data(row.get('DATA_LIBERACAO'))
            d_liq = sin_converte_data(row.get('DATA_LIQUIDACAO'))

            if pd.notna(d_sin):
                if (pd.notna(d_avi) and d_sin > d_avi) or (pd.notna(d_vis) and d_sin > d_vis) or \
                   (pd.notna(d_lib) and d_sin > d_lib) or (pd.notna(d_liq) and d_sin > d_liq):
                    error_message.append("Data do sinistro deve ser anterior a todas as outras")
        except Exception as e: error_message.append(f"Erro ao comparar datas: {e}")

        df.at[i, 'Error'] = ', '.join(error_message) if error_message else 'Erros não localizados'
        set_progress_safe(root, canvas, ((i + 1) / max(total, 1)) * 100)

    return df

# -------------------------------------------------------------
# Orquestração
# -------------------------------------------------------------
CATEGORIAS = [
    'Garantias para Locação (Não utilizado)',
    'Seguros Imobiliários',
    'Faturamento Educacional',
    'Sinistro Educacional'
]

def processar_arquivo(root, canvas, status_label, categoria, input_path, output_path, btn):
    try:
        update_status_safe(root, status_label, "Lendo arquivo Excel...", "#F1C40F")
        df = pd.read_excel(input_path)

        update_status_safe(root, status_label, "Analisando regras de diagnóstico...", "#00E5FF")
        
        if categoria == 'Garantias para Locação (Não utilizado)': df_out = pipeline_garantias_locacao(df, root, canvas)
        elif categoria == 'Seguros Imobiliários': df_out = pipeline_seguros_imobiliarios(df, root, canvas)
        elif categoria == 'Faturamento Educacional': df_out = pipeline_faturamento_educacional(df, root, canvas)
        elif categoria == 'Sinistro Educacional': df_out = pipeline_sinistro_educacional(df, root, canvas)
        else: raise ValueError('Categoria inválida')

        update_status_safe(root, status_label, "Salvando arquivo...", "#F1C40F")
        df_out.to_excel(output_path, index=False)
        
        update_status_safe(root, status_label, "Aplicando design no Excel...", "#9B59B6")
        formatar_planilha_saida(output_path)
        
        update_status_safe(root, status_label, "✔ Análise Concluída com Sucesso!", "#00E676")
        show_info_safe(root, 'Sucesso', 'Arquivo verificado, gerado e formatado com sucesso!')
        
    except Exception as e:
        update_status_safe(root, status_label, f"✖ Erro Crítico: {str(e)[:40]}...", "#E74C3C")
        show_error_safe(root, 'Erro', f'Ocorreu um erro: {e}')
    finally:
        root.after(0, lambda: btn.configure(state="normal"))

def iniciar_processo(root, combo, canvas, status_label, btn):
    categoria = combo.get()
    if not categoria:
        messagebox.showwarning('Aviso', 'Selecione uma categoria primeiro!')
        return

    input_path = filedialog.askopenfilename(title='Selecione a base de dados', filetypes=[('Excel', '*.xlsx;*.xls')])
    if not input_path: return

    output_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx')], title='Salvar Relatório de Diagnóstico')
    if not output_path: return

    btn.configure(state="disabled")
    desenhar_foguete(canvas, 0)
    update_status_safe(root, status_label, "Iniciando propulsores...", "#00E5FF")

    threading.Thread(target=processar_arquivo, args=(root, canvas, status_label, categoria, input_path, output_path, btn), daemon=True).start()

# -------------------------------------------------------------
# GUI PRINCIPAL NEXT-GEN (CustomTkinter)
# -------------------------------------------------------------
def criar_interface():
    app = ctk.CTk()
    app.title("Sistema de Diagnóstico - Core")
    app.geometry("600x450")
    app.resizable(False, False)

    # Frame Central
    main_frame = ctk.CTkFrame(app, corner_radius=15, fg_color="#1E1E2E")
    main_frame.pack(pady=30, padx=30, fill="both", expand=True)

    # Título
    titulo = ctk.CTkLabel(main_frame, text="CENTRAL DE DIAGNÓSTICO", font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"), text_color="#00E5FF")
    titulo.pack(pady=(25, 15))

    # Seleção de Categoria
    lbl_modulo = ctk.CTkLabel(main_frame, text="Módulo Operacional:", font=ctk.CTkFont(family="Segoe UI", size=12))
    lbl_modulo.pack(anchor="w", padx=40)

    combo = ctk.CTkOptionMenu(main_frame, values=CATEGORIAS, 
                              fg_color="#2b2d42", button_color="#00E5FF", button_hover_color="#00b4cc",
                              dropdown_fg_color="#1E1E2E", font=("Segoe UI", 13), width=460)
    combo.pack(pady=(5, 20), padx=40)
    combo.set('Seguros Imobiliários')

    # Foguete Canvas (Integrado ao visual dark)
    lbl_status = ctk.CTkLabel(main_frame, text="Aguardando inicialização do sistema...", font=ctk.CTkFont(family="Segoe UI", size=12, slant="italic"), text_color="#8d99ae")
    lbl_status.pack(pady=(10, 0))

    canvas = tk.Canvas(main_frame, height=50, bg="#1E1E2E", highlightthickness=0)
    canvas.pack(fill="x", padx=40, pady=(5, 20))
    desenhar_foguete(canvas, 0)

    # Botão de Ação
    btn_executar = ctk.CTkButton(main_frame, text="INICIAR VARREDURA", 
                                 font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
                                 fg_color="#00E5FF", text_color="#121212", hover_color="#00b4cc",
                                 height=45, corner_radius=8)
    
    btn_executar.configure(command=lambda: iniciar_processo(app, combo, canvas, lbl_status, btn_executar))
    btn_executar.pack(fill="x", padx=40, pady=(0, 20))

    app.mainloop()

if __name__ == '__main__':
    criar_interface()
